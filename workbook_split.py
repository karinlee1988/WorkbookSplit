#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time : 2022/7/10 11:38
# @Author : karinlee
# @FileName : workbook_split.py
# @Software : PyCharm
# @Blog : https://blog.csdn.net/weixin_43972976
# @github : https://github.com/karinlee1988/
# @gitee : https://gitee.com/karinlee/
# @Personal website : https://karinlee.cn/


import openpyxl

class WorksheetSplitByColumn(object):
    """
    用于按列的内容拆分excel工作薄中的一个工作表，分别另存为多个独立的工作薄/打印
    """

    def __init__(self,filepath,column_index:int,title_index:int):

        """

        :param filepath: excel xlsx文件路径
        :param column_index: 按哪列的内容进行拆分（A列为1，B列为2,C列为3...）
        :param title_index: 表头的行数（从1开始，即表头有几行就填写几行）

        """
        # 获取文件路径
        self.filepath = filepath
        # 获取作为条件拆分的列号
        self.column_index = column_index
        # 获取表头行数
        self.title_index =title_index

    def split(self, column_key: str):
        """
        拆分表格并分别另存为

        在待拆分表格中，采取按列关键字删除的方式来进行拆分，避免出现格式错乱
        这样，待拆分表格先行用excel打开设置好格式，删除就不会错乱
        但删除操作耗时较大。

        :param column_key:  关键字用于按列拆分
        :return:
        """
        # 读取待拆分的源数据表（为了不改变原有格式，每次都读取一次比较稳妥。拆分等于就是在原表上将不要的行删除掉，格式还是原表的格式）
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        # 根据传入的列关键字，若列中与传入的关键字不同，则删除。循环完毕后，剩下就等效于拆分后的数据了
        # 删除操作会改变行对应的行号，所以从最大行开始遍历，倒过来循环。
        for row in range(ws.max_row + 1, self.title_index, -1):
            if ws.cell(row=row, column=5).value != column_key:
                ws.delete_rows(row)
        wb.save(f'{column_key}.xlsx')


if __name__ == '__main__':
    workbook_path = '托收单附件明细待拆分打印.xlsx'
    app = WorksheetSplitByColumn(workbook_path,0,3)
    app.split('英德市白沙镇中心小学')






